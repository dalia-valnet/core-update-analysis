"""
Write an excel file for a Google core update analysis
Get URL Clicks, Impressions, Average Position, URL CTR, Count of Landing Pages, and the Difference in each of these metrics from 'Before Period' to 'After Period'
 Summarize these metrics by Content Type, Primary Category
 [**Later include by Pub/Repub, Year published, etc.]

 python3 analysis.py [arguments]

 arguments in this order
    prevDate: dates you exported the data for written as one string, 
    curDate: , 
    intent: file name from ga export landing pages+intent.
    contentType: file name from ga export of landing pages with content type. Needs to be .csv file, 
    primCat: same but for category,
    metricsPrev: file name from sc export. Needs to be .csv file,
    metricsCur: same,
    filename: name of the file you want to dump the results in 

    make sure all files listed in arguments are in the same folder as analysis.py

    e.g. python analysis.py sept22_oct5 oct6_oct14 ga_export_content.csv ga_export_cat.csv sc_before.csv sc_after.csv hc_core_update_Oct_analysis.xlsx
"""

 

import pandas as pd
#numpy used for datemod, datepub etc
#import numpy as np
#import matplotlib.pyplot as plt

#import seaborn as sns
#import os
import sys
#from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
import re


def extract_post(url):
    post_type = str(re.search(r'(/($|[a-zA-Z-]+)(/|$))', url).group())
    if post_type[-1:]!='/':
        post_type = post_type+'/'
    return post_type


def loadData(
    intent: str,
    contentType: str, 
    primCat: str,
    metricsPrev: str,
    metricsCur: str
    ):

    # FUNCTION returns dataframe of merged GA and SC data - Landing pages with content type and primary category + metrics before and after update

    df_intent = pd.read_csv(intent, encoding='latin-1',usecols=[0,1])
    df_intent.columns = ['Landing Page', 'Intent']
    
    df_contentType = pd.read_csv(contentType, encoding='latin-1',usecols=[0,1])
    df_contentType.columns = ['Landing Page', 'Content Type']

    df_primCat = pd.read_csv(primCat, encoding='latin-1',usecols=[0,1])
    df_primCat.columns = ['Landing Page', 'Primary Category']
    #df_primCat.columns = ['Landing Page', 'Hub']

    # keep only primary category
    #*** Add if len >=2 ...  
    #df_primCat['Primary Category'] = df_primCat['Primary Category'].str.split('|').str[1]

    #combine the previous and current period data
    df_prev = pd.read_csv(metricsPrev, encoding='latin-1')
    df_prev['Period'] = 'Before update'
    df_prev.columns = ['Landing Page']+df_prev.columns[1:].to_list() 

    df_cur = pd.read_csv(metricsCur, encoding='latin-1')
    df_cur['Period'] = 'After update'
    df_cur.columns = ['Landing Page']+df_cur.columns[1:].to_list() 

    metrics = pd.concat([df_prev, df_cur])

    # ** make sure using Full landing page in all df (do it on Looker before export)
    #metrics['Landing Page'] = metrics['Landing Page'].str.replace(f'^https://(www\.)?(.*).com', '', regex=True)

    #merge GA and SC tables
    df = pd.merge(metrics, df_intent, on='Landing Page', how='left')
    df = pd.merge(df, df_contentType, on='Landing Page', how='left')
    df = pd.merge(df, df_primCat, on='Landing Page', how='left')

    #### FOR CARBUZZ ###
    #df = metrics
    #df['Post Type'] = df['Landing Page'].apply(extract_post)
    #df = pd.merge(df, df_contentType, on='Landing Page', how='left')


    # Will group the rows on Landing Page, keep the different period data seperate. (lik pivot table by Landing page)
    # But since GA data contains duplicates (some pages have more than one content type or category so they appear more than once), we'll only keep the first occurence of the page
    # ****FIX : Dont just take first, take max...?
    df = df.groupby(['Landing Page', 'Period']).first()
    df = df.reset_index()

    # **LATER add here any more cleaning or manipulation want to do to the data. Also add datePub, dateMod df if want them later and create df_repub

    return df

def calculate(
    data,
    groups: list,   
):
    # Dataframe - summarize by whats specified in groups
        # make sure is last in the 'groups' list
    # Groupby already by default sorts groups (landing page, then period) ascending
    df = data.groupby(groups).agg({
        'Url Clicks': 'sum',
        'Impressions': 'sum',
        'Average Position': 'mean',
        'Landing Page': 'count'
    })

    # Calculations
    # If we're grouping by more than just Period, then the calculations are performed differently
    # **LATER if want to group by more than just Period and one other dimension, write the calculations to handle those cases

    if(len(groups) == 1):
        df['URL CTR'] = df['Url Clicks']/df['Impressions']
        df['Change URL CTR'] = df['URL CTR'] - df['URL CTR'].shift(-1)
        df['% Change URL Clicks'] = (df['Url Clicks'] - df['Url Clicks'].shift(-1))/(df['Url Clicks']).shift(-1)
        df['% Change Impressions'] = (df['Impressions'] - df['Impressions'].shift(-1))/(df['Impressions']).shift(-1)
        df['Position Difference'] = df['Average Position'] - df['Average Position'].shift(-1)

    if(len(groups) == 2):
        df['URL CTR'] = df['Url Clicks']/df['Impressions']
        df['Change URL CTR'] = df['URL CTR'] - df['URL CTR'].groupby(groups[0]).shift(-1)
        df['% Change URL Clicks'] = (df['Url Clicks'] - df['Url Clicks'].groupby(groups[0]).shift(-1))/(df['Url Clicks']).groupby(groups[0]).shift(-1)
        df['% Change Impressions'] = (df['Impressions'] - df['Impressions'].groupby(groups[0]).shift(-1))/(df['Impressions']).groupby(groups[0]).shift(-1)
        df['Position Difference'] = df['Average Position'] - df['Average Position'].groupby(groups[0]).shift(-1)

    #set the order of the columns and rows
    df = df[[
        'Landing Page', 
        'Url Clicks', 
        '% Change URL Clicks', 
        'Impressions', 
        '% Change Impressions', 
        'Average Position', 
        'Position Difference', 
        'URL CTR', 
        'Change URL CTR']]

    return df


def formatting(
    sheet,
    row,
    nb_col
):
    #sheet.sheet_view.showGridLines = False
    red_fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    max_row = sheet.max_row
    max_col = sheet.max_column
            
    for col_idx in range(1, max_col + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)

        # Set all the column widths to 25
        sheet.column_dimensions[col_letter].width = 25

        is_index_col = col_idx <= nb_col
                
        for row_idx in range(row+2, max_row + 1):
            cell = sheet[f'{col_letter}{row_idx}']
            if cell.value is None or cell.value == '' or is_index_col:
                continue
                    
            col_name = sheet[f'{col_letter}{row+1}'].value
                
            # Convert cell value to float if possible and apply formatting based on column name
            try:
                if isinstance(cell.value, str) and '%' in cell.value:
                    cell_value = float(cell.value.strip('%')) / 100  # Convert percentage to float
                else:
                    cell_value = float(cell.value)
                        
                # Apply number format and conditional formatting based on column name
                if col_name in ['Url Clicks', 'Impressions']:
                    cell.number_format = '#,##0'
                        
                elif col_name in ['% Change URL Clicks', '% Change Impressions', 'Change URL CTR']:
                    cell.number_format = '0.00%'
                    if cell_value > 0:
                        cell.fill = green_fill
                    elif cell_value < 0:
                        cell.fill = red_fill
                            
                elif col_name in ['Average Position']:
                    cell.number_format = '0.00'
                        
                elif col_name in ['Position Difference']:
                    cell.number_format = '0.00'
                    if cell_value > 0:
                        cell.fill = red_fill
                    elif cell_value < 0:
                        cell.fill = green_fill
                        
                elif col_name == 'URL CTR':
                    cell.number_format = '0.00%'
                            
            except ValueError:
                continue  # Skip non-convertible (non-numeric) cells
                        

def writeExcel(
    data,
    tables: dict,
    file: str,
    prevDate,
    curDate
):
    with pd.ExcelWriter(file, engine = 'openpyxl') as writer:

        #loop through the tables, writing each one to a new sheet in file
        for sheet_name, table in tables.items():
            
            # on the Overall sheet, we want to also write down the dates, so we will dump df 3 rows lower than other sheets
            row = 0
            if sheet_name == 'Overall':
                row = 3

            #to_excel by default sets NAN to empty string values, and also writes row names (index = True)
            table.to_excel(writer, sheet_name=sheet_name, startrow=row)
            sheet = writer.sheets[sheet_name]
            if sheet_name == 'Overall':
                sheet['A1'] = 'Before Period'
                sheet['A2'] = 'After Period'
                sheet['B1'] = prevDate
                sheet['B2'] = curDate

            formatting(writer.sheets[sheet_name], row, len(table.index.names))

        data.to_excel(writer, sheet_name = 'Data')


def analyze(
    args
    ):
    
    #### Want 3 sheets: 1.Overall metrics  2. metrics by Content Type   3. metrics by Primary Category
    ### Bonus sheet: write all merged data to a 4th sheet

    prevDate = args[0] 
    curDate = args[1] 
    intent = args[2]
    contentType = args[3] 
    primCat = args[4]
    metricsPrev = args[5]
    metricsCur = args[6]
    filename = args[7]

    # read in all the data from the different files. Make sure data files have correct columns and in the right order.
    data = loadData(intent, contentType, primCat, metricsPrev, metricsCur)

    # Using the compiled data, calculate the metrics that we want for the analysis: % Change in clicks and impressions, change in Position, CTR and change in CTR
    # Different groupings: first Overall summary, then group by content type, then group by primary category
    df_overall = calculate(data, ['Period'])
    df_intent = calculate(data ,['Intent', 'Period'])
    df_content = calculate(data, ['Content Type', 'Period'])
    df_primCat = calculate(data, ['Primary Category', 'Period'])
    ### FOR CB ###
    #df_content = calculate(data, ['Post Type', 'Period'])
    #df_primCat = calculate(data, ['Hub', 'Period'])

    tables = {
        'Overall': df_overall,
        'Intent': df_intent,
        'Content Type': df_content,
        'Primary Category': df_primCat
    }

    # write the calculations to a new excel file. Each sheet is a different summary grouping
    writeExcel(data, tables, filename, prevDate, curDate)



if __name__ == '__main__':
    args = sys.argv
    analyze(args[1:])






